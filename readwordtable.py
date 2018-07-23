from docx import Document
from openpyxl import Workbook

def word_to_excel(version):
    filepath = '/Users/pengyi/PycharmProjects/pengyi/data/work_split/'
    filename = version + "_需求任务拆解.docx"
    workbook = Workbook()
    booksheet = workbook.active
    booksheet.append([0,1,2,3,4])

    f = Document(filepath + filename)
    tables = f.tables
    n = 1
    for table in tables:
        len_row = len(table.rows)
        n += 1
        print(len_row)
        print(table)
        for i in range(1, len_row):
            list = []
            model = ''
            work = ''
            owner = ''
            spend_time = ''
            end_time = ''

            model = table.cell(i,0).text
            work = table.cell(i, 1).text
            owner = table.cell(i, 2).text
            spend_time = table.cell(i, 3).text
            end_time = table.cell(i, 4).text

            list.append(model.strip())
            list.append(work.strip())
            list.append(owner.strip())
            list.append(spend_time.strip())
            list.append(end_time.strip())

            booksheet.append(list)

    workbook.save('/Users/pengyi/PycharmProjects/pengyi/data/work_split/' + version + '_需求任务拆解.xlsx')

if __name__ == '__main__':
    version = 'SDP2_v0.3'
    word_to_excel(version)